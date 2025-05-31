import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tksheet
import os
import shutil
import csv
import json
import math
import pandas as pd
from io import StringIO
from PIL import Image, ImageTk
from datetime import datetime, timedelta

CONFIG_FILE = "app_config.json"


def get_default_config():
    return {
        "general": {
            "image_dir": "",
            "output_dir": "",
            "csv_file": "",
            "last_export_csv_filename": ""
        },
        "format_config": {},
        "columns_config": {},
        "datetime_format": "%Y-%m-%d %H:%M:%S.%f",
        "png_datetime_format": {
            "date_format": "%d%m%Y",
            "time_format": "%H%M%S"
        },
        "mandatory_export_columns": [
            "Line", "Point", "NodeCode", "Index",
            "Aslaid Time", "Recovered Time",
            "Deployed by ROV", "Recovered by ROV",
            "DeployedComments", "RecoveredComments"
        ],
        "defaults": {
            "datetime_format": "%Y-%m-%d %H:%M:%S.%f",
            "png_date_format": "%d%m%Y",
            "png_time_format": "%H%M%S",
            "mandatory_export_columns": [
                "Line", "Point", "NodeCode", "Index",
                "Aslaid Time", "Recovered Time",
                "Deployed by ROV", "Recovered by ROV",
                "DeployedComments", "RecoveredComments"
            ]
        }
    }


def load_app_config():
    config = get_default_config()
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                user_config = json.load(f)
            for k, v in user_config.items():
                config[k] = v
        except Exception:
            pass
    return config


def save_app_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)

def extract_date_time_from_filename(filename, format_config, date_fmt, time_fmt):
    """
    Extract datetime from filename using format_config (start/len for Date/Time) and supplied formats.
    Returns a pandas.Timestamp or pd.NaT.
    """
    fname = os.path.splitext(os.path.basename(filename))[0]
    date_info = format_config.get('Date', {})
    time_info = format_config.get('Time', {})
    try:
        d_start = int(date_info.get('start', 0))
        d_len = int(date_info.get('length', 0))
        t_start = int(time_info.get('start', 0))
        t_len = int(time_info.get('length', 0))
        date_str = fname[d_start:d_start+d_len]
        time_str = fname[t_start:t_start+t_len]
        dt = pd.to_datetime(date_str + time_str, format=(date_fmt + time_fmt), errors='coerce')
        return dt
    except Exception:
        return pd.NaT

def strip_rov_whitespace_columns(export_df):
    """
    Strip outside whitespace for all columns in export_df whose name contains 'ROV' (case-insensitive).
    Modifies export_df in place.
    """
    rov_cols = [col for col in export_df.columns if 'rov' in col.lower()]
    for col in rov_cols:
        export_df[col] = export_df[col].astype(str).str.strip()

def create_export_df(events_df, deployment_df, recovery_df):
    """
    Outer merge events_df and deployment_df on ['Line', 'Point', 'Index'].
    Create 'Node Name' as concatenation of (str) Line, Point, Index.
    'Node Name' is placed as the first column in the resulting DataFrame.
    Returns the merged export_df.
    """
    key_cols = ['Line', 'Point', 'Index']
    # Ensure input DataFrames have key columns as string for matching/concat
    for df in (events_df, deployment_df):
        if df is not None and not df.empty:
            for col in key_cols:
                if col in df.columns:
                    df[col] = df[col].astype(str)

    # Outer merge
    if events_df is not None and not events_df.empty and deployment_df is not None and not deployment_df.empty:
        export_df = pd.merge(events_df, deployment_df, on=key_cols, how='outer', suffixes=('_event', '_deploy'))
    elif events_df is not None and not events_df.empty:
        export_df = events_df.copy()
    elif deployment_df is not None and not deployment_df.empty:
        export_df = deployment_df.copy()
    else:
        # No data to export
        return pd.DataFrame()

    # Insert Node Name column
    export_df.insert(
        0,
        "Node Name",
        export_df["Line"].astype(str) + export_df["Point"].astype(str) + export_df["Index"].astype(str)
    )

    # Clean up ROV columns
    rov_cols = [col for col in export_df.columns if 'rov' in col.lower()]
    for col in rov_cols:
        export_df[col] = export_df[col].astype(str).str.strip()

    for column in ['Bumper','ROV','Datetime','filename']:
        if column in export_df.columns:
            export_df.rename(columns={column: column+'_dep'},inplace=True)

##    # Deduplicate by 'Node Name', keeping latest 'Datetime'
##    if "Node Name" in export_df.columns and "Datetime" in export_df.columns:
##        export_df = export_df.sort_values(["Node Name", "Datetime"], ascending=[True, False])
##        export_df = export_df.drop_duplicates(subset=["Node Name"], keep="first")
##        export_df = export_df.sort_values("Datetime", ascending=False).reset_index(drop=True)

    # Merge with recovery_df if provided and not empty
    if recovery_df is not None and not recovery_df.empty:
        for col in key_cols:
            if col in recovery_df.columns:
                recovery_df[col] = recovery_df[col].astype(str)
        export_df = pd.merge(export_df, recovery_df, on=key_cols, how='outer', suffixes=('_dep', '_rec'))

    for column in ['Bumper','ROV','Datetime','filename']:
        if column in export_df.columns:
            export_df.rename(columns={column: column+'_rec'},inplace=True)

##    # Deduplicate by 'Node Name' for recovery, keeping latest 'Datetime_rec'
##    if "Node Name" in export_df.columns and "Datetime_rec" in export_df.columns:
##        export_df = export_df.sort_values(["Node Name", "Datetime_rec"], ascending=[True, False])
##        export_df = export_df.drop_duplicates(subset=["Node Name"], keep="first")
##        export_df = export_df.sort_values("Datetime_rec", ascending=False).reset_index(drop=True)

    # Clean NodeCode and Bumper columns
    if "NodeCode" in export_df.columns:
        export_df["NodeCode"] = export_df["NodeCode"].astype(str).str.strip().str.replace(" ", "", regex=False)

    bumper_cols = [col for col in export_df.columns if 'bumper' in col.lower()]
    for col in bumper_cols:
        export_df[col] = export_df[col].astype(str).str.lstrip("0")

    # Final column order (only keep those present)
    cols_list = ['Node Name', 'NodeCode', 'Bumper_dep', 'Bumper_rec', 'Deployed by ROV', 'ROV_dep',
                  'Recovered by ROV', 'ROV_rec', 'Aslaid Time','Datetime_dep', 'Recovered Time','Datetime_rec',
                  'filename_dep','filename_rec','DeployedComments','RecoveredComments']
    export_df = export_df.reindex(columns=cols_list)

    # Sort by Node Name at the end
    if "Node Name" in export_df.columns:
        export_df = export_df.sort_values("Node Name").reset_index(drop=True)

    return export_df


class FilenameFormatDialog(ttk.LabelFrame):
    def __init__(
        self,
        master,
        app_config,
        save_all_config,
        get_deployment_png_filenames,
        on_format_change,
        on_update_export_data,  # <-- add this
        *args,
        **kwargs
    ):
        super().__init__(master, text="PNG Filename Format Definition", *args, **kwargs)
        self.app_config = app_config
        self.save_all_config = save_all_config
        self.get_deployment_png_filenames = get_deployment_png_filenames
        self.on_format_change = on_format_change
        self.format_items = ["Line", "Point", "Index", "Bumper", "Date", "Time", "ROV"]
        self.format_vars = {}
        self.format_entries = {}
        self.sample_labels = []
        self.active_field = (None, None)
        self.last_png_samples = []
        self.png_date_format_var = tk.StringVar()
        self.png_time_format_var = tk.StringVar()
        self.example_label = None
        self.initialized = False
        self.on_update_export_data = on_update_export_data
        self.init_widgets()
        self.load_last_format_config()
        self.load_png_datetime_format()
        self.update_samples()
        self.initialized = True
        

    def init_widgets(self):
        container = ttk.Frame(self)
        container.grid(row=0, column=0, sticky="nsew")
        self.columnconfigure(0, weight=1)
        container.columnconfigure(0, weight=1)
        container.columnconfigure(1, weight=0)

        left_frame = ttk.Frame(container)
        left_frame.grid(row=0, column=0, sticky="nsew")
        right_frame = ttk.Frame(container)
        right_frame.grid(row=0, column=1, sticky="nsew", padx=(15, 0), pady=5)

        sample_frame = ttk.Frame(left_frame)
        sample_frame.grid(row=0, column=0, columnspan=20, sticky="ew", pady=(4, 8))
        self.sample_labels = []
        for i in range(3):
            lbl = tk.Text(
                sample_frame, width=50, height=1, font=("TkDefaultFont", 10),
                relief="flat", background="white", highlightthickness=0
            )
            lbl.grid(row=i, column=0, sticky="ew", pady=0)
            lbl.config(state="disabled")
            self.sample_labels.append(lbl)

        table_frame = ttk.Frame(left_frame)
        table_frame.grid(row=1, column=0, sticky="w", padx=2, pady=4)

        ttk.Label(table_frame, text="").grid(row=0, column=0, padx=3, pady=2, sticky="w")
        for idx, item in enumerate(self.format_items):
            ttk.Label(table_frame, text=item).grid(row=0, column=idx+1, padx=10, pady=2, sticky="w")

        ttk.Label(table_frame, text="Start").grid(row=1, column=0, padx=3, pady=2, sticky="e")
        for idx, item in enumerate(self.format_items):
            start_var = tk.StringVar()
            self.format_vars.setdefault(item, {})["start"] = start_var
            start_entry = tk.Entry(table_frame, textvariable=start_var, width=6)
            start_entry.grid(row=1, column=idx+1, padx=(0, 10), pady=1)
            start_entry.bind("<FocusIn>", lambda e, it=item: self._activate_field(it, "start"))
            start_entry.bind("<FocusOut>", lambda e, it=item: self._deactivate_field(it, "start"))
            start_entry.bind("<KeyRelease>", lambda e, it=item: self._validate_and_update(it, "start"))
            self.format_entries.setdefault(item, {})["start"] = start_entry

        ttk.Label(table_frame, text="Len").grid(row=2, column=0, padx=3, pady=2, sticky="e")
        for idx, item in enumerate(self.format_items):
            length_var = tk.StringVar()
            self.format_vars[item]["length"] = length_var
            length_entry = tk.Entry(table_frame, textvariable=length_var, width=6)
            length_entry.grid(row=2, column=idx+1, padx=(0, 10), pady=1)
            length_entry.bind("<FocusIn>", lambda e, it=item: self._activate_field(it, "length"))
            length_entry.bind("<FocusOut>", lambda e, it=item: self._deactivate_field(it, "length"))
            length_entry.bind("<KeyRelease>", lambda e, it=item: self._validate_and_update(it, "length"))
            self.format_entries.setdefault(item, {})["length"] = length_entry

        png_dt_frame = ttk.LabelFrame(right_frame, text="Filename Date/Time Format")
        png_dt_frame.pack(fill="x", padx=3, pady=5, anchor="n")
        ttk.Label(png_dt_frame, text="Date format:").grid(row=0, column=0, sticky="w", padx=(8,4), pady=2)
        date_entry = ttk.Entry(png_dt_frame, textvariable=self.png_date_format_var, width=14)
        date_entry.grid(row=0, column=1, padx=(0,4), pady=2)
        date_entry.bind("<FocusOut>", lambda e: self.save_png_datetime_format())
        date_entry.bind("<Return>", lambda e: self.save_png_datetime_format())
        ttk.Label(png_dt_frame, text="Time format:").grid(row=1, column=0, sticky="w", padx=(8,4), pady=2)
        time_entry = ttk.Entry(png_dt_frame, textvariable=self.png_time_format_var, width=14)
        time_entry.grid(row=1, column=1, padx=(0,4), pady=2)
        time_entry.bind("<FocusOut>", lambda e: self.save_png_datetime_format())
        time_entry.bind("<Return>", lambda e: self.save_png_datetime_format())
        ttk.Button(png_dt_frame, text="Reset default", command=self.reset_png_datetime_format).grid(row=2, column=0, columnspan=2, pady=4)

        self.example_label = ttk.Label(png_dt_frame, text="", foreground="blue", anchor="w", justify="left")
        self.example_label.grid(row=3, column=0, columnspan=2, sticky="w", padx=8, pady=(6,2))

        for item in self.format_items:
            self._validate_and_update(item, "start")
            self._validate_and_update(item, "length")

        # ttk.Button(left_frame, text="Save Format", command=self.save_current_format_config).grid(row=3, column=0, pady=8)
        ttk.Button(left_frame, text="Update Export Data", command=self.on_update_export_data).grid(row=3, column=0, pady=8)

    def _activate_field(self, item, field):
        self.active_field = (item, field)
        self.update_samples()
        self._validate_and_update(item, field)

    def _deactivate_field(self, item, field):
        self.active_field = (None, None)
        self.update_samples()
        self._validate_and_update(item, field)

    def _validate_and_update(self, item, field):
        entry = self.format_entries[item][field]
        value = self.format_vars[item][field].get()
        if value == "" or (value.isdigit() and int(value) >= 0):
            entry.configure(background="white")
        else:
            entry.configure(background="#ffcccc")
        self.update_samples()
        if self.initialized and self.on_format_change:
            self.on_format_change()

    def save_current_format_config(self):
        format_config = {}
        for item in self.format_items:
            start = self.format_vars[item]["start"].get()
            length = self.format_vars[item]["length"].get()
            format_config[item] = {"start": start, "length": length}
        self.app_config["format_config"] = format_config
        self.save_all_config()
        if self.on_format_change:
            self.on_format_change()

    def load_last_format_config(self):
        format_config = self.app_config.get("format_config", {})
        for item in self.format_items:
            vals = format_config.get(item, {})
            self.format_vars[item]["start"].set(vals.get("start", ""))
            self.format_vars[item]["length"].set(vals.get("length", ""))

    def load_png_datetime_format(self):
        png_dt = self.app_config.get("png_datetime_format", {})
        defaults = self.app_config.get("defaults", {})
        self.png_date_format_var.set(png_dt.get("date_format", defaults.get("png_date_format", "%d%m%Y")))
        self.png_time_format_var.set(png_dt.get("time_format", defaults.get("png_time_format", "%H%M%S")))
        self.update_example_conversion()

    def save_png_datetime_format(self):
        defaults = self.app_config.get("defaults", {})
        fmt = {
            "date_format": self.png_date_format_var.get() or defaults.get("png_date_format", "%d%m%Y"),
            "time_format": self.png_time_format_var.get() or defaults.get("png_time_format", "%H%M%S")
        }
        self.png_date_format_var.set(fmt["date_format"])
        self.png_time_format_var.set(fmt["time_format"])
        self.app_config["png_datetime_format"] = fmt
        self.save_all_config()
        self.update_example_conversion()
        if self.on_format_change:
            self.on_format_change()

    def reset_png_datetime_format(self):
        defaults = self.app_config.get("defaults", {})
        self.png_date_format_var.set(defaults.get("png_date_format", "%d%m%Y"))
        self.png_time_format_var.set(defaults.get("png_time_format", "%H%M%S"))
        self.save_png_datetime_format()

    def get_current_format_config(self):
        return {
            item: {
                "start": self.format_vars[item]["start"].get(),
                "length": self.format_vars[item]["length"].get()
            }
            for item in self.format_items
        }

    def update_example_conversion(self):
        # Use a real (or sample) filename from user context
        filenames = self.get_deployment_png_filenames() if self.get_deployment_png_filenames else []
        sample_fname = os.path.splitext(os.path.basename(filenames[0]))[0] if filenames else "20240529_143501_..."
        format_config = self.get_current_format_config()
        date_fmt = self.png_date_format_var.get() or self.app_config.get("defaults", {}).get("png_date_format", "%d%m%Y")
        time_fmt = self.png_time_format_var.get() or self.app_config.get("defaults", {}).get("png_time_format", "%H%M%S")
        # NOTE: you must provide extract_date_time_from_filename externally
        dt = extract_date_time_from_filename(sample_fname, format_config, date_fmt, time_fmt)
        now = datetime.now()
        ten_years = timedelta(days=365 * 10)
        error_message = ""
        if pd.isnull(dt):
            error_message = f"❌ Error: Could not extract date/time (NaT) from '{sample_fname}' with current settings."
        elif abs(dt - now) > ten_years:
            error_message = f"❌ Error: Parsed date/time {dt.strftime('%Y-%m-%d %H:%M:%S')} is more than 10 years from today."
        if error_message:
            self.example_label.config(text=error_message, background="#ffe5e5", foreground="red")
        else:
            example = f"Example: '{sample_fname}' → {dt.strftime('%Y-%m-%d %H:%M:%S')}"
            self.example_label.config(text=example, background="", foreground="blue")

    def update_samples(self):
        filenames = self.get_deployment_png_filenames()[:3] if self.get_deployment_png_filenames else []
        self.last_png_samples = filenames
        format_config = self.get_current_format_config()
        active_item, active_field = self.active_field
        for i, lbl in enumerate(self.sample_labels):
            lbl.config(state="normal")
            lbl.delete("1.0", tk.END)
            if i < len(filenames):
                fname = os.path.splitext(os.path.basename(filenames[i]))[0]
                lbl.insert("1.0", fname)
                if active_item is not None:
                    start = format_config[active_item].get("start", "")
                    length = format_config[active_item].get("length", "")
                    try:
                        s = int(start)
                        l = int(length)
                        if s >= 0 and l > 0 and s < len(fname):
                            e = min(s + l, len(fname))
                            lbl.tag_add("highlight", f"1.{s}", f"1.{e}")
                    except Exception:
                        pass
                lbl.tag_configure("highlight", background="#ffff00")
            lbl.config(state="disabled")
        self.update_example_conversion()

class CSVColumnDialog(ttk.Frame):
    """
    Adapted for unified config:
    - columns_config must be a reference to self.app_config["columns_config"]
    - updates are persisted by calling parent.save_all_config()
    """
    def __init__(self, master, columns, columns_config, on_change, save_all_config=None, app_config=None, *args, **kwargs):
        super().__init__(master, *args, **kwargs)
        self.app_config = app_config
        self.columns = columns
        self.columns_config = columns_config
        self.on_change = on_change
        self.save_all_config = save_all_config
        self.vars = {}

        # This must be provided or imported elsewhere
        self.mandatory_columns = set(self.app_config.get("mandatory_export_columns", []))

        # Create a canvas for horizontal scrolling
        canvas = tk.Canvas(self, borderwidth=0, height=40)
        frame = ttk.Frame(canvas)
        h_scroll = ttk.Scrollbar(self, orient="horizontal", command=canvas.xview)
        canvas.configure(xscrollcommand=h_scroll.set)

        h_scroll.pack(side="bottom", fill="x")
        canvas.pack(side="top", fill="x", expand=True)
        canvas.create_window((0, 0), window=frame, anchor='nw')

        self.checkbuttons = {}
        for i, col in enumerate(self.columns):
            is_mandatory = col in self.mandatory_columns
            var = tk.BooleanVar(value=True if is_mandatory else columns_config.get(col, True))
            state = "disabled" if is_mandatory else "normal"
            cb = ttk.Checkbutton(frame, text=col, variable=var, command=self._on_change, state=state)
            cb.grid(row=0, column=i, padx=3, sticky="w")
            self.vars[col] = var
            self.checkbuttons[col] = cb
            # Ensure config is always True for mandatory columns
            if is_mandatory:
                self.columns_config[col] = True

        def on_configure(event):
            canvas.config(scrollregion=canvas.bbox("all"))
        frame.bind("<Configure>", on_configure)

        self.frame = frame
        self.canvas = canvas

        # "Only Mandatory" tickbox
        self.only_mandatory_var = tk.BooleanVar(value=False)
        only_mandatory_cb = ttk.Checkbutton(self, text="Only Mandatory", variable=self.only_mandatory_var, command=self.on_only_mandatory)
        only_mandatory_cb.pack(side="top", anchor="w", pady=(2, 0))

    def on_only_mandatory(self):
        only = self.only_mandatory_var.get()
        for col, var in self.vars.items():
            is_mandatory = col in self.mandatory_columns
            if not is_mandatory:
                var.set(False if only else self.columns_config.get(col, True))
                self.columns_config[col] = var.get()
        self._on_change()

    def _on_change(self):
        for col, var in self.vars.items():
            # Mandatory columns always True
            if col in self.mandatory_columns:
                self.columns_config[col] = True
            else:
                self.columns_config[col] = var.get()
        if self.save_all_config:
            self.save_all_config()
        self.on_change()

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.app_config = load_app_config()
        self.save_all_config = lambda: save_app_config(self.app_config)
        self.format_config = self.app_config.setdefault("format_config", {})
        self.png_datetime_format = self.app_config.setdefault("png_datetime_format", {
            "date_format": self.app_config["defaults"]["png_date_format"],
            "time_format": self.app_config["defaults"]["png_time_format"]
        })
        self.columns_config = self.app_config.setdefault("columns_config", {})
        self.datetime_format = self.app_config.setdefault("datetime_format", self.app_config["defaults"]["datetime_format"])
        self.mandatory_export_columns = self.app_config.setdefault("mandatory_export_columns", self.app_config["defaults"]["mandatory_export_columns"])
        self.general = self.app_config.setdefault("general", {})
        self.csv_data = []
        self.events_df = None
        self.deployment_df = pd.DataFrame()
        self.recovery_df = pd.DataFrame()
        self.export_df = pd.DataFrame()
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True)
        self.tab_images = ttk.Frame(self.notebook)
        self.tab_csv = ttk.Frame(self.notebook)
        self.tab_process = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_images, text="PNG Files")
        self.notebook.add(self.tab_csv, text="CSV File")
        self.notebook.add(self.tab_process, text="Process & Export")
        self.image_viewer_window = None
        self.image_viewer_image_label = None
        self.image_viewer_img = None
        self.image_viewer_files = []
        self.image_viewer_index = 0
        self.image_viewer_source = None
        self.init_tab_images()
        self.init_tab_csv()
        self.init_tab_process()
        self.load_last_choices()
                # --- Excel Export Section ---
        self.excel_filename_var = tk.StringVar()
        self.excel_col_map_vars = {}  # {excel_col: tk.StringVar}
        self.excel_cols = []
        self.excel_file_chosen = False

    # ---------- PNG FORMAT PARSER UTILS ----------
    def parse_filename_fields(self, fname, format_config):
        basename = os.path.splitext(os.path.basename(fname))[0]
        result = {}
        for field in ["Line", "Point", "Index", "Bumper", "Date", "Time", "ROV"]:
            conf = format_config.get(field, {})
            start = conf.get("start")
            length = conf.get("length")
            if start is None or length is None or not start.isdigit() or not length.isdigit():
                result[field] = None
                continue
            s = int(start)
            l = int(length)
            result[field] = basename[s:s+l] if s < len(basename) else ""
        return result

    def filelist_to_dataframe(self, files, format_config, png_datetime_format):
        rows = []
        for f in files:
            info = self.parse_filename_fields(f, format_config)
            if info is not None:
                row = {**info}
                row['filename'] = os.path.basename(f)
                rows.append(row)
        df = pd.DataFrame(rows)
        if "Date" in df.columns and "Time" in df.columns:
            date_fmt = png_datetime_format.get("date_format", self.app_config["defaults"].get("png_date_format", "%d%m%Y"))
            time_fmt = png_datetime_format.get("time_format", self.app_config["defaults"].get("png_time_format", "%H%M%S"))
            df["Date"] = df["Date"].fillna("")
            df["Time"] = df["Time"].fillna("")
            dt_str = df["Date"].astype(str) + df["Time"].astype(str)
            mask = (df["Date"].str.len() == len("20240529")) & (df["Time"].str.len() == len("143501"))
            try:
                dt_vals = pd.to_datetime(dt_str.where(mask), format=date_fmt + time_fmt, errors="coerce")
            except Exception:
                dt_vals = pd.NaT
            df["Datetime"] = dt_vals
            df = df.drop(columns=["Date", "Time"])
        return df

    def try_update_deployment_recovery_dataframes(self):
        image_dir = self.general.get("image_dir", "")
        format_dialog_ready = hasattr(self, "format_dialog") and getattr(self.format_dialog, "get_current_format_config", None)
        filename_format_defined = format_dialog_ready and self.format_dialog.get_current_format_config()
        if image_dir and os.path.isdir(image_dir):
            self.populate_deployment_recovery_lists(image_dir)
            if filename_format_defined:
                self.update_deployment_recovery_dataframes()

    def update_deployment_recovery_dataframes(self):
        if not hasattr(self, "format_dialog"):
            return
        format_config = self.format_dialog.get_current_format_config()

        def find_png_files(subfolder_name, base_folder):
            results = []
            for root, dirs, _ in os.walk(base_folder):
                if subfolder_name in dirs:
                    target_dir = os.path.join(root, subfolder_name)
                    for dp, dn, filenames in os.walk(target_dir):
                        results.extend(
                            os.path.join(dp, f)
                            for f in filenames
                            if f.lower().endswith('.png')
                        )
                    break
            return results

        folder = self.image_dir_var.get()
        deployment_files, recovery_files = [], []
        if folder and os.path.isdir(folder):
            deployment_files = find_png_files("Deployment", folder)
            recovery_files = find_png_files("Recovery", folder)

        format_config = self.format_dialog.get_current_format_config()
        png_datetime_format = {
            "date_format": self.format_dialog.png_date_format_var.get(),
            "time_format": self.format_dialog.png_time_format_var.get(),
        }
        self.deployment_df = self.filelist_to_dataframe(deployment_files, format_config, png_datetime_format)
        self.recovery_df = self.filelist_to_dataframe(recovery_files, format_config, png_datetime_format)

    # ---------- IMAGES TAB ----------
    def init_tab_images(self):
        frm = ttk.Frame(self.tab_images)
        frm.pack(fill="both", expand=True, padx=10, pady=10)
        self.image_dir_var = tk.StringVar(value=self.general.get("image_dir", ""))

        self.nav_dir_frame = ttk.Frame(self.tab_images)
        self.nav_dir_frame.pack(fill="x", pady=(10, 0))
        self.nav_image_dir_var = tk.StringVar(value=self.general.get("nav_image_dir", ""))
        ttk.Label(self.nav_dir_frame, text="Nav Image Folder:").pack(side="left")
        ttk.Entry(self.nav_dir_frame, textvariable=self.nav_image_dir_var, width=60, state="readonly").pack(side="left", padx=5)
        ttk.Button(self.nav_dir_frame, text="Choose...", command=self.choose_nav_image_folder).pack(side="left")
        ttk.Button(self.nav_dir_frame, text="Copy Nav Images to Pictures Folder", command=self.copy_nav_images_to_pictures).pack(side="left", padx=(10, 0))
        self.copy_progress_var = tk.DoubleVar(value=0)
        self.copy_progress = ttk.Progressbar(
            self.nav_dir_frame, orient="horizontal", length=200, mode="determinate",
            variable=self.copy_progress_var, maximum=100
        )
        self.copy_progress.pack(side="left", padx=(10, 0))

        self.copied_message_var = tk.StringVar(value="")
        self.copied_message_label = ttk.Label(self.nav_dir_frame, textvariable=self.copied_message_var)
        self.copied_message_label.pack(side="left", padx=(10, 0))

        dir_frame = ttk.Frame(frm)
        dir_frame.pack(fill="x")
        ttk.Label(dir_frame, text="Pictures Folder:").pack(side="left")
        ttk.Entry(dir_frame, textvariable=self.image_dir_var, width=60, state="readonly").pack(side="left", padx=5)
        ttk.Button(dir_frame, text="Choose...", command=self.choose_image_folder).pack(side="left")
        ttk.Button(dir_frame, text="Refresh", command=self.refresh_png_views).pack(side="left", padx=(10, 0))

        bottom_frame = ttk.Frame(frm)
        bottom_frame.pack(fill="both", expand=True, pady=10)

        self.deployment_frame = ttk.LabelFrame(bottom_frame, text="Deployment")
        self.deployment_frame.pack(side="left", fill="both", expand=True, padx=(0,5))
        self.deployment_listbox = tk.Listbox(self.deployment_frame, xscrollcommand=lambda *args: self.deploy_hscroll.set(*args), yscrollcommand=lambda *args: self.deploy_vscroll.set(*args))
        self.deployment_listbox.pack(side="left", fill="both", expand=True)
        self.deploy_vscroll = ttk.Scrollbar(self.deployment_frame, orient="vertical", command=self.deployment_listbox.yview)
        self.deploy_vscroll.pack(side="right", fill="y")
        self.deploy_hscroll = ttk.Scrollbar(self.deployment_frame, orient="horizontal", command=self.deployment_listbox.xview)
        self.deploy_hscroll.pack(fill="x")
        self.deployment_listbox.config(yscrollcommand=self.deploy_vscroll.set, xscrollcommand=self.deploy_hscroll.set)
        self.deployment_listbox.bind('<<ListboxSelect>>', lambda e: self.open_selected_image('deployment'))

        self.recovery_frame = ttk.LabelFrame(bottom_frame, text="Recovery")
        self.recovery_frame.pack(side="left", fill="both", expand=True, padx=(5,0))
        self.recovery_listbox = tk.Listbox(self.recovery_frame, xscrollcommand=lambda *args: self.recovery_hscroll.set(*args), yscrollcommand=lambda *args: self.recovery_vscroll.set(*args))
        self.recovery_listbox.pack(side="left", fill="both", expand=True)
        self.recovery_vscroll = ttk.Scrollbar(self.recovery_frame, orient="vertical", command=self.recovery_listbox.yview)
        self.recovery_vscroll.pack(side="right", fill="y")
        self.recovery_hscroll = ttk.Scrollbar(self.recovery_frame, orient="horizontal", command=self.recovery_listbox.xview)
        self.recovery_hscroll.pack(fill="x")
        self.recovery_listbox.config(yscrollcommand=self.recovery_vscroll.set, xscrollcommand=self.recovery_hscroll.set)
        self.recovery_listbox.bind('<<ListboxSelect>>', lambda e: self.open_selected_image('recovery'))

        self.try_update_deployment_recovery_dataframes()

    def populate_deployment_recovery_lists(self, folder):
        deployment_path = None
        recovery_path = None
        for root, dirs, files in os.walk(folder):
            for d in dirs:
                if d == "Deployment" and deployment_path is None:
                    deployment_path = os.path.join(root, d)
                if d == "Recovery" and recovery_path is None:
                    recovery_path = os.path.join(root, d)
        self.deployment_listbox.delete(0, tk.END)
        if deployment_path and os.path.isdir(deployment_path):
            pngs = self.list_png_files_recursive(deployment_path)
            if pngs:
                for fname in pngs:
                    self.deployment_listbox.insert(tk.END, fname)
                maxlen = max((len(f) for f in pngs), default=40)
                self.deployment_listbox.config(width=max(40, min(maxlen, 200)))
            else:
                self.deployment_listbox.insert(tk.END, "No PNG files found in 'Deployment'")
                self.deployment_listbox.config(width=40)
        else:
            self.deployment_listbox.insert(tk.END, "'Deployment' subdirectory not found")
            self.deployment_listbox.config(width=40)

        self.recovery_listbox.delete(0, tk.END)
        if recovery_path and os.path.isdir(recovery_path):
            pngs = self.list_png_files_recursive(recovery_path)
            if pngs:
                for fname in pngs:
                    self.recovery_listbox.insert(tk.END, fname)
                maxlen = max((len(f) for f in pngs), default=40)
                self.recovery_listbox.config(width=max(40, min(maxlen, 200)))
            else:
                self.recovery_listbox.insert(tk.END, "No PNG files found in 'Recovery'")
                self.recovery_listbox.config(width=40)
        else:
            self.recovery_listbox.insert(tk.END, "'Recovery' subdirectory not found")
            self.recovery_listbox.config(width=40)

        if hasattr(self, "format_dialog"):
            self.format_dialog.update_samples()
        self.update_deployment_recovery_dataframes()

    def choose_image_folder(self):
        initial_dir = self.general.get("image_dir", "")
        folder = filedialog.askdirectory(
            title="Select Pictures Folder",
            initialdir=initial_dir if os.path.isdir(initial_dir) else None
        )
        if folder:
            self.image_dir_var.set(folder)
            self.general["image_dir"] = folder
            self.save_all_config()
            self.populate_deployment_recovery_lists(folder)

    def choose_nav_image_folder(self):
        initial_dir = self.general.get("nav_image_dir", "")
        folder = filedialog.askdirectory(
            title="Select Nav Image Folder",
            initialdir=initial_dir if os.path.isdir(initial_dir) else None
        )
        if folder:
            self.nav_image_dir_var.set(folder)
            self.general["nav_image_dir"] = folder
            self.save_all_config()

    def copy_nav_images_to_pictures(self):
        nav_folder = self.nav_image_dir_var.get()
        pictures_folder = self.image_dir_var.get()
        if not (nav_folder and os.path.isdir(nav_folder) and pictures_folder and os.path.isdir(pictures_folder)):
            messagebox.showerror("Error", "Please select valid Nav and Pictures folders first.")
            return

        def update_progress(current, total):
            percent = 100 * current / total if total else 0
            self.copy_progress_var.set(percent)
            self.copy_progress.update_idletasks()
            self.copied_message_var.set("")  # Clear previous message

        # Reset progress
        self.copy_progress_var.set(0)
        self.copy_progress.update_idletasks()

        # Optionally run in a thread to avoid UI freezing for lots of files
        import threading
        def run_copy():
            total_files_copied = self.copy_nav_to_pictures(nav_folder, pictures_folder, progress_callback=update_progress)
            self.copy_progress_var.set(100)
            self.try_update_deployment_recovery_dataframes()
            self.copied_message_var.set(f"Copied {total_files_copied} new PNG files.")
            
        threading.Thread(target=run_copy, daemon=True).start()

    def copy_nav_to_pictures(self, nav_folder, pictures_folder, progress_callback=None):
        deployment_base = os.path.join(pictures_folder, "Deployment")
        recovery_base = os.path.join(pictures_folder, "Recovery")

        # Collect all files to copy for progress reporting
        file_list = []
        for line_name in os.listdir(nav_folder):
            line_path = os.path.join(nav_folder, line_name)
            if not os.path.isdir(line_path):
                continue
            for subfolder in os.listdir(line_path):
                subfolder_path = os.path.join(line_path, subfolder)
                if not os.path.isdir(subfolder_path):
                    continue
                subfolder_lower = subfolder.lower()
                if subfolder_lower.startswith('deploy'):
                    dest_base = deployment_base
                elif subfolder_lower.startswith('recover'):
                    dest_base = recovery_base
                else:
                    continue
                for fname in os.listdir(subfolder_path):
                    if fname.lower().endswith('.png'):
                        file_list.append((subfolder, line_name, subfolder_path, fname, dest_base))

        total_files = len(file_list)
        copied_count = 0

        for idx, (subfolder, line_name, src_subfolder, fname, dest_base) in enumerate(file_list):
            dest_line_folder = os.path.join(dest_base, line_name)
            os.makedirs(dest_line_folder, exist_ok=True)
            src_file = os.path.join(src_subfolder, fname)
            dest_file = os.path.join(dest_line_folder, fname)

            # Skip if file exists and creation time matches
            if os.path.exists(dest_file):
                # File already exists with same creation time, skip
                if progress_callback:
                    progress_callback(idx + 1, total_files)
                continue

            shutil.copy2(src_file, dest_file)
            copied_count += 1
            if progress_callback:
                progress_callback(idx + 1, total_files)
        return copied_count
        # Example usage:
        # nav_folder = self.nav_image_dir_var.get()
        # pictures_folder = self.image_dir_var.get()
        # copy_nav_to_pictures(nav_folder, pictures_folder)

    def refresh_png_views(self):
        folder = self.image_dir_var.get()
        if folder and os.path.isdir(folder):
            self.populate_deployment_recovery_lists(folder)
        else:
            messagebox.showinfo("Info", "No valid pictures folder selected.")

    def list_png_files_recursive(self, root_folder):
        png_files = []
        for root, _, files in os.walk(root_folder):
            for file in files:
                if file.lower().endswith('.png'):
                    relpath = os.path.relpath(os.path.join(root, file), root_folder)
                    png_files.append(relpath)
        return png_files

    def get_deployment_png_filenames(self):
        folder = self.image_dir_var.get()
        if not (folder and os.path.isdir(folder)):
            return []
        deployment_path = None
        for root, dirs, files in os.walk(folder):
            for d in dirs:
                if d == "Deployment":
                    deployment_path = os.path.join(root, d)
                    break
            if deployment_path:
                break
        if deployment_path and os.path.isdir(deployment_path):
            return self.list_png_files_recursive(deployment_path)
        else:
            return []

    def open_selected_image(self, source):
        if source == "deployment":
            listbox = self.deployment_listbox
            folder = self.image_dir_var.get()
            subfolder = "Deployment"
        else:
            listbox = self.recovery_listbox
            folder = self.image_dir_var.get()
            subfolder = "Recovery"
        if not folder or not os.path.isdir(folder):
            return
        files = []
        dir_path = None
        for root, dirs, _ in os.walk(folder):
            if subfolder in dirs:
                dir_path = os.path.join(root, subfolder)
                break
        if not dir_path or not os.path.isdir(dir_path):
            return
        for root, _, fs in os.walk(dir_path):
            for f in fs:
                if f.lower().endswith('.png'):
                    files.append(os.path.join(root, f))
        files.sort()
        sel = listbox.curselection()
        if not sel:
            return
        selected_idx = sel[0]
        if selected_idx >= len(files):
            return
        full_path = files[selected_idx]
        self.open_image_viewer(files, selected_idx, source)

    def open_image_viewer(self, files, index, source):
        if self.image_viewer_window is not None and tk.Toplevel.winfo_exists(self.image_viewer_window):
            self.image_viewer_files = files
            self.image_viewer_index = index
            self.image_viewer_source = source
            self.update_image_viewer()
            self.image_viewer_window.lift()
            return

        self.image_viewer_files = files
        self.image_viewer_index = index
        self.image_viewer_source = source

        win = tk.Toplevel(self)
        win.title("Image Viewer")
        win.geometry("800x600")
        self.image_viewer_window = win

        lbl = tk.Label(win)
        lbl.pack(expand=True, fill="both")
        self.image_viewer_image_label = lbl

        nav = ttk.Frame(win)
        nav.pack(fill="x")
        btn_back = ttk.Button(nav, text="Back", command=self.image_viewer_back)
        btn_back.pack(side="left", padx=10, pady=5)
        btn_forward = ttk.Button(nav, text="Forward", command=self.image_viewer_forward)
        btn_forward.pack(side="right", padx=10, pady=5)

        self.update_image_viewer()

        def on_close():
            self.image_viewer_window = None
            self.image_viewer_image_label = None
            self.image_viewer_img = None
            self.image_viewer_files = []
            self.image_viewer_index = 0
            self.image_viewer_source = None
            win.destroy()
        win.protocol("WM_DELETE_WINDOW", on_close)

    def update_image_viewer(self):
        if not self.image_viewer_files or self.image_viewer_index < 0 or self.image_viewer_index >= len(self.image_viewer_files):
            return
        filepath = self.image_viewer_files[self.image_viewer_index]
        try:
            img = Image.open(filepath)
            w, h = img.size
            max_w, max_h = 760, 520
            ratio = min(max_w / w, max_h / h, 1.0)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.Resampling.LANCZOS)
            self.image_viewer_img = ImageTk.PhotoImage(img)
            self.image_viewer_image_label.config(image=self.image_viewer_img)
            self.image_viewer_window.title(f"Image Viewer - {os.path.basename(filepath)} ({self.image_viewer_index+1}/{len(self.image_viewer_files)})")
        except Exception as e:
            self.image_viewer_image_label.config(text=f"Could not open image:\n{filepath}\n{e}", image="")
            self.image_viewer_img = None

    def image_viewer_back(self):
        if self.image_viewer_index > 0:
            self.image_viewer_index -= 1
            self.update_image_viewer()
            self._select_in_listbox(self.image_viewer_source, self.image_viewer_index)

    def image_viewer_forward(self):
        if self.image_viewer_files and self.image_viewer_index < len(self.image_viewer_files) - 1:
            self.image_viewer_index += 1
            self.update_image_viewer()
            self._select_in_listbox(self.image_viewer_source, self.image_viewer_index)

    def _select_in_listbox(self, source, index):
        if source == "deployment":
            lb = self.deployment_listbox
        else:
            lb = self.recovery_listbox
        lb.selection_clear(0, tk.END)
        lb.selection_set(index)
        lb.see(index)

    def init_tab_csv(self):
        frm = ttk.Frame(self.tab_csv)
        frm.pack(fill="both", expand=True, padx=10, pady=10)
        self.csv_file_var = tk.StringVar(value=self.general.get("csv_file", ""))

        file_frame = ttk.Frame(frm)
        file_frame.pack(fill="x")
        ttk.Label(file_frame, text="CSV File:").pack(side="left")
        ttk.Entry(file_frame, textvariable=self.csv_file_var, width=60, state="readonly").pack(side="left", padx=5)
        ttk.Button(file_frame, text="Choose...", command=self.choose_csv_file).pack(side="left")

        self.csv_text_frame = ttk.Frame(frm)
        self.csv_text_frame.pack(fill="x", pady=(10, 2))
        self.csv_text = tk.Text(self.csv_text_frame, wrap="none", height=6)
        self.csv_text.pack(side="left", fill="x", expand=True)
        text_vscroll = ttk.Scrollbar(self.csv_text_frame, orient="vertical", command=self.csv_text.yview)
        text_vscroll.pack(side="right", fill="y")
        self.csv_text.config(yscrollcommand=text_vscroll.set)
        text_hscroll = ttk.Scrollbar(frm, orient="horizontal", command=self.csv_text.xview)
        text_hscroll.pack(fill="x")
        self.csv_text.config(xscrollcommand=text_hscroll.set)
        self.csv_text.config(state="disabled")

        self.col_dialog_frame = ttk.Frame(frm)
        self.col_dialog_frame.pack(fill="x", pady=(2, 4))
        self.csv_col_dialog = None

        self.datetime_frame = ttk.LabelFrame(frm, text="Datetime Format")
        self.datetime_frame.pack(fill="x", pady=(2, 6), padx=3)
        ttk.Label(self.datetime_frame, text="Datetime format:").pack(side="left", padx=(5, 2))
        self.datetime_format_var = tk.StringVar(value=self.datetime_format)
        self.datetime_entry = ttk.Entry(self.datetime_frame, textvariable=self.datetime_format_var, width=35)
        self.datetime_entry.pack(side="left", padx=(0, 5))
        ttk.Button(self.datetime_frame, text="Reset default", command=self.reset_datetime_format).pack(side="left", padx=3)
        self.datetime_entry.bind("<FocusOut>", lambda e: self.save_datetime_format_from_entry())
        self.datetime_entry.bind("<Return>", lambda e: self.save_datetime_format_from_entry())

        self.events_df_frame = ttk.LabelFrame(frm, text="Events dataframe")
        self.events_df_frame.pack(fill="both", expand=True, pady=(4, 2))
        self.events_df_sheet = None  # Will become a tksheet.Sheet instance


        csv_file = self.general.get("csv_file", "")
        if csv_file and os.path.isfile(csv_file):
            self.load_csv_file(csv_file)

    def save_datetime_format_from_entry(self):
        fmt = self.datetime_format_var.get()
        if not fmt:
            fmt = DEFAULT_DATETIME_FORMAT
            self.datetime_format_var.set(fmt)
        self.app_config["datetime_format"] = fmt
        self.save_all_config()
        self.datetime_format = fmt

    def reset_datetime_format(self):
        self.datetime_format_var.set(DEFAULT_DATETIME_FORMAT)
        save_datetime_format(DEFAULT_DATETIME_FORMAT)
        self.datetime_format = DEFAULT_DATETIME_FORMAT

    def choose_csv_file(self):
        file = filedialog.askopenfilename(title="Select CSV File", filetypes=[("CSV Files", "*.csv")])
        if file:
            self.csv_file_var.set(file)
            self.general["csv_file"] = file
            self.save_all_config()
            self.load_csv_file(file)

    def load_csv_file(self, file):
        self.csv_data.clear()
        try:
            with open(file, "r", encoding="utf-8") as txtfile:
                text = txtfile.read()
            if hasattr(self, "csv_text"):
                self.csv_text.config(state="normal")
                self.csv_text.delete(1.0, tk.END)
                self.csv_text.insert(tk.END, text)
                self.csv_text.config(state="disabled")
        except Exception as e:
            if hasattr(self, "csv_text"):
                self.csv_text.config(state="normal")
                self.csv_text.delete(1.0, tk.END)
                self.csv_text.insert(tk.END, f"Failed to load file as text: {e}")
                self.csv_text.config(state="disabled")
            return

        # Ensure datetime_format is up to date and saved
        if hasattr(self, "save_datetime_format_from_entry"):
            self.save_datetime_format_from_entry()
        dt_format = getattr(self, "datetime_format_var", None)
        if dt_format:
            dt_format = self.datetime_format_var.get()
        if not dt_format:
            dt_format = self.app_config.get("datetime_format", self.app_config["defaults"]["datetime_format"])

        try:
            with open(file, "r", encoding="utf-8") as txtfile:
                lines = txtfile.readlines()
            valid_lines = [line for line in lines if line.strip() and not line.strip().startswith("#")]
            if not valid_lines:
                messagebox.showerror("CSV Error", "No valid lines in selected file.")
                return
            first_valid = valid_lines[0]
            col_names = [c.strip() for c in next(csv.reader([first_valid]))]
            data_lines = valid_lines[1:]
            csv_str = first_valid + "".join(data_lines)
            df = pd.read_csv(
                StringIO(csv_str),
                comment="#",
                skip_blank_lines=True,
                dtype=str,
                dayfirst=False,
                keep_default_na=False,
            )
            # Only parse dates for columns that exist
            parse_dates = [col for col in df.columns if "Time" in col or "time" in col]
            for col in parse_dates:
                df[col] = pd.to_datetime(df[col], errors="coerce", format=dt_format)

            self.events_df = df
            # Robust column config using app_config
            columns_config = self.app_config.setdefault("columns_config", {})
            for col in self.events_df.columns:
                if col not in columns_config:
                    columns_config[col] = True
            self.save_all_config()
            # If you have a CSVColumnDialog, update or recreate it as needed:
            if hasattr(self, "csv_col_dialog") and self.csv_col_dialog:
                self.csv_col_dialog.destroy()
            if hasattr(self, "col_dialog_frame"):
                self.csv_col_dialog = CSVColumnDialog(
                        self.col_dialog_frame,
                        list(self.events_df.columns),
                        self.app_config["columns_config"],
                        self.update_dataframe_view,
                        save_all_config=self.save_all_config,
                        app_config=self.app_config
                    )
                self.csv_col_dialog.pack(fill="x")
            self.columns_config = columns_config
            if hasattr(self, "update_dataframe_view"):
                self.update_dataframe_view()
        except Exception as e:
            messagebox.showerror("CSV Error", f"Failed to import file: {e}")

    def update_dataframe_view(self):
        if self.events_df is None:
            return
        selected_cols = [col for col, show in self.columns_config.items() if show and col in self.events_df.columns]
        if not selected_cols:
            reduced_df = self.events_df.iloc[:, []]
        else:
            reduced_df = self.events_df[selected_cols]
        # Destroy old tksheet if present
        if hasattr(self, "events_df_sheet") and self.events_df_sheet:
            self.events_df_sheet.destroy()
        headers = list(reduced_df.columns)
        data = reduced_df.values.tolist() if not reduced_df.empty else []
        self.events_df_sheet = tksheet.Sheet(
            self.events_df_frame,
            data=data,
            headers=headers,
            height=600,  # pixels; adjust as needed
            show_x_scrollbar=True,
            show_y_scrollbar=True,
        )
        self.events_df_sheet.enable_bindings((
            "single_select", "row_select", "column_select", "arrowkeys"
        ))
        self.events_df_sheet.pack(fill="both", expand=True)

    # PROCESS & EXPORT TAB

    def init_tab_process(self):
        frm = ttk.Frame(self.tab_process)
        frm.pack(fill="both", expand=True, padx=10, pady=10)
        self.output_dir_var = tk.StringVar(value=self.general.get("output_dir", ""))
        out_frame = ttk.Frame(frm)
        out_frame.pack(fill="x", pady=(0, 5))
        ttk.Label(out_frame, text="Output Directory:").pack(side="left")
        ttk.Entry(out_frame, textvariable=self.output_dir_var, width=60, state="readonly").pack(side="left", padx=5)
        ttk.Button(out_frame, text="Choose...", command=self.choose_output_dir).pack(side="left")

        self.format_dialog = FilenameFormatDialog(
            frm,
            app_config=self.app_config,
            save_all_config=self.save_all_config,
            get_deployment_png_filenames=self.get_deployment_png_filenames,
            on_format_change=self.update_deployment_recovery_dataframes,
            on_update_export_data=self.update_export_data
        )
        self.format_dialog.pack(fill="x", pady=(5, 10), padx=5)

        self.try_update_deployment_recovery_dataframes()

        self.excel_filename_var = tk.StringVar()
        self.btn_frame = ttk.Frame(frm)
        self.btn_frame.pack(fill="both", expand=True, pady=(0, 10))

        # Use grid to place two frames side by side, each half width
        self.btn_frame.columnconfigure(0, weight=1)
        self.btn_frame.columnconfigure(1, weight=1)
        self.btn_frame.rowconfigure(0, weight=1)

        process_csv_frame = ttk.Frame(self.btn_frame)
        process_csv_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 12), pady=0)

        self.excel_frame = ttk.Frame(self.btn_frame)
        self.excel_frame.grid(row=0, column=1, sticky="nsew", padx=(0, 0), pady=0)

        self.excel_mapping_frame = ttk.Frame(self.excel_frame)
        self.excel_mapping_frame.pack(fill="x", expand=False)

        self.excel_filename_label = ttk.Label(
            self.excel_mapping_frame,
            text="",  # We'll set the text dynamically
            foreground="blue"
        )
        self.excel_filename_label.pack(anchor="w", pady=2, fill="x", expand=True)

        excel_btns_frame = ttk.Frame(self.excel_frame)
        excel_btns_frame.pack(side="bottom", fill="x", pady=(4, 0))

        ttk.Button(excel_btns_frame, text="Export Fixed Excel Format", command=self.export_excel).pack(side="left", padx=2)
        ttk.Button(excel_btns_frame, text="Choose Excel File to Update", command=self.choose_excel_file_to_update).pack(side="left", padx=2)
        ttk.Button(excel_btns_frame, text="Update", command=self.update_existing_excel_from_ui).pack(side="left", padx=2) 

        self.bind("<Configure>", self._update_excel_filename_wraplength)

        # --- CSV Export Section ---
        self.csv_columns_vars = {}
        self.csv_check_frame = ttk.LabelFrame(process_csv_frame, text="Columns to export (CSV)")
        self.csv_check_frame.pack(fill="both", expand=True, padx=3, pady=3)
        ttk.Button(process_csv_frame, text="Export CSV", command=self.export_csv).pack(side="bottom", pady=(4, 0))

        self.export_frame = ttk.Frame(frm)
        self.export_frame.pack(fill="both", expand=True)

        self.update_idletasks()
        half_width = self.btn_frame.winfo_width() // 2
        process_csv_frame.config(width=half_width)
        self.excel_frame.config(width=half_width)

        # Export frame for the tksheet widget
        self.export_frame = tk.Frame(frm)
        self.export_frame.pack(fill="both", expand=True)

        self.process_status = tk.StringVar()
        ttk.Label(frm, textvariable=self.process_status).pack()

        
            
    def _update_excel_filename_wraplength(self, event=None):
        # Use excel_frame's width, not the root's!
        if hasattr(self, "excel_frame"):
            width = self.excel_frame.winfo_width()
            wrap = max(width - 20, 100)
            self.excel_filename_label.configure(wraplength=wrap)

    def populate_csv_column_checkboxes(self):
        # Remove previous
        for widget in self.csv_check_frame.winfo_children():
            widget.destroy()
        self.csv_columns_vars.clear()
        export_cols = list(self.export_df.columns)

        # Load only columns that are still available in export_df
        last_cols = self.general.get("csv_export_columns", export_cols)
        # Only tick those that exist now
        for i, col in enumerate(export_cols):
            var = tk.BooleanVar(value=(col in last_cols))
            chk = ttk.Checkbutton(self.csv_check_frame, text=col, variable=var)
            chk.grid(row=i, column=0, sticky="w")
            self.csv_columns_vars[col] = var

    def update_excel_mapping_ui(self):
        for widget in self.excel_mapping_frame.winfo_children():
            widget.destroy()
        if not self.excel_file_chosen or not self.excel_cols:
            return
        # File name label
        self.excel_filename_label = ttk.Label(
            self.excel_mapping_frame,
            text=self.excel_filename_var.get(),
            foreground="blue"
        )
        self.excel_filename_label.pack(anchor="w", pady=2, fill="x", expand=True)
        self._update_excel_filename_wraplength()  # Make sure wraplength is correct after re-creating label
        
        # Get the last mapping from config
        last_map = self.general.get("last_update_excel_mapping", {})
        # Only preselect if the mapped value exists in export_df columns
        export_df_cols = list(self.export_df.columns)
        for i, excel_col in enumerate(self.excel_cols):
            ttk.Label(self.excel_mapping_frame, text=excel_col).pack(anchor="w", pady=(3 if i > 0 else 6,0))
            mapped_col = last_map.get(excel_col, "")
            if mapped_col not in export_df_cols:
                mapped_col = ""
            var = tk.StringVar(value=mapped_col)
            values = [""] + export_df_cols
            box = ttk.Combobox(self.excel_mapping_frame, textvariable=var, values=values, width=28, state="readonly")
            box.pack(anchor="w", padx=12)
            self.excel_col_map_vars[excel_col] = var

    def save_excel_mapping_from_ui(self):
        # Save current UI mapping to config
        current_map = {}
        for excel_col, var in self.excel_col_map_vars.items():
            df_col = var.get()
            if df_col:  # Only save non-empty mappings
                current_map[excel_col] = df_col
        self.general["last_update_excel_mapping"] = current_map
        self.save_all_config()  # If you want to persist it right away


    def update_export_data(self):
        mandatory_columns = self.app_config.get("mandatory_export_columns", [])
        filtered_events_df = self.events_df[[col for col in mandatory_columns if col in self.events_df.columns]].copy()
        self.export_df = create_export_df(filtered_events_df, self.deployment_df, self.recovery_df)
        self.show_export_df_with_cell_highlight()
        self.populate_csv_column_checkboxes()

    def choose_excel_file_to_update(self):
        filename = filedialog.askopenfilename(
            title="Select Excel file to update",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=os.path.basename(self.app_config.get("last_update_excel_file", "")),
            initialdir=os.path.dirname(self.app_config.get("last_update_excel_file", "")) or ""
        )
        if not filename:
            return
        self.excel_filename_var.set(filename)
        self.app_config["last_update_excel_file"] = filename
        self.save_all_config()
        # Load columns from Excel file
        from openpyxl import load_workbook
        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]
        self.excel_cols = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        # Restore last mapping if present:
        last_map = self.app_config.get("last_update_excel_mapping", {})
        self.excel_col_map_vars = {col: tk.StringVar(value=last_map.get(col, "")) for col in self.excel_cols}
        self.excel_file_chosen = True
        self.update_excel_mapping_ui()

    def show_export_df_with_cell_highlight(self):
        # Clear previous widgets
        for widget in self.export_frame.winfo_children():
            widget.destroy()

        df = self.export_df

        # Create tksheet Sheet widget
        sheet = tksheet.Sheet(
            self.export_frame,
            data=df.values.tolist(),
            headers=list(df.columns),
            show_x_scrollbar=True,
            show_y_scrollbar=True, height = 600
        )
        
        sheet.pack(fill="both", expand=True)

        # ----------- Existing ROV/Deploy highlight logic -----------
        for i, row in df.iterrows():
            deployed_by_rov = row.get('Deployed by ROV', '')
            rov_dep = row.get('ROV_dep', '')
            if (pd.notnull(deployed_by_rov)
                and str(deployed_by_rov).strip()
                and deployed_by_rov != rov_dep):
                col1 = df.columns.get_loc('Deployed by ROV')
                col2 = df.columns.get_loc('ROV_dep')
                sheet.highlight_cells(row=i, column=col1, bg="#ffcc99")  # Orange
                sheet.highlight_cells(row=i, column=col2, bg="#ffcc99")  # Orange

        # ----------- New: Highlight duplicate Node Name cells -----------
        if "Node Name" in df.columns:
            node_name_col = df.columns.get_loc("Node Name")
            node_name_counts = df["Node Name"].value_counts()
            duplicates = set(node_name_counts[node_name_counts > 1].index)
            for i, val in enumerate(df["Node Name"]):
                if val in duplicates and pd.notnull(val) and str(val).strip():
                    sheet.highlight_cells(row=i, column=node_name_col, bg="#ffff00")  # Yellow

        # ----------- Highlight Bumper_dep if not same as NodeCode -----------
        if "Bumper_dep" in df.columns and "NodeCode" in df.columns:
            bumper_dep_col = df.columns.get_loc("Bumper_dep")
            for i, row in df.iterrows():
                bumper_dep_val = row.get("Bumper_dep", '')
                nodecode_val = row.get("NodeCode", '')
                if bumper_dep_val != nodecode_val:
                    sheet.highlight_cells(row=i, column=bumper_dep_col, bg="#ff0000")  # Red

        # ----------- Highlight Bumper_rec if present and not same as NodeCode -----------
        if "Bumper_rec" in df.columns and "NodeCode" in df.columns:
            bumper_rec_col = df.columns.get_loc("Bumper_rec")
            for i, row in df.iterrows():
                bumper_rec_val = row.get("Bumper_rec", '')
                nodecode_val = row.get("NodeCode", '')
                if (
                    pd.notnull(bumper_rec_val)
                    and str(bumper_rec_val).strip()
                    and bumper_rec_val != nodecode_val
                ):
                    sheet.highlight_cells(row=i, column=bumper_dep_col, bg="#ff0000")  # Red

        # ----------- NEW: Highlight Datetime_dep if abs diff to Aslaid Time > 15 min -----------
        if "Aslaid Time" in df.columns and "Datetime_dep" in df.columns:
            aslaid_series = pd.to_datetime(df["Aslaid Time"], errors="coerce")
            datetime_dep_series = pd.to_datetime(df["Datetime_dep"], errors="coerce")
            col_dep = df.columns.get_loc("Datetime_dep")
            for i, (a, d) in enumerate(zip(aslaid_series, datetime_dep_series)):
                if pd.notnull(a) and pd.notnull(d) and abs((a - d).total_seconds()) > 900:
                    sheet.highlight_cells(row=i, column=col_dep, bg="#A9D08E")  # Green

        # ----------- NEW: Highlight Datetime_rec if abs diff to Recovered Time > 15 min -----------
        if "Recovered Time" in df.columns and "Datetime_rec" in df.columns:
            recovered_series = pd.to_datetime(df["Recovered Time"], errors="coerce")
            datetime_rec_series = pd.to_datetime(df["Datetime_rec"], errors="coerce")
            col_rec = df.columns.get_loc("Datetime_rec")
            for i, (r, d) in enumerate(zip(recovered_series, datetime_rec_series)):
                if pd.notnull(r) and pd.notnull(d) and abs((r - d).total_seconds()) > 900:
                    sheet.highlight_cells(row=i, column=col_rec, bg="#A9D08E")  # Green

        sheet.enable_bindings((
            "single_select", "row_select", "column_select", "drag_select",
            "row_drag_and_drop", "column_drag_and_drop", "arrowkeys",
            "right_click_popup_menu", "rc_select", "copy", "cut", "paste",
            "delete", "undo", "edit_cell", "column_width_resize"
        ))

    def export_csv(self):
        # Save selected columns to config
        selected_cols = [col for col, var in self.csv_columns_vars.items() if var.get()]
        self.general["csv_export_columns"] = selected_cols
        self.save_all_config()

        if getattr(self, "export_df", None) is None or self.export_df.empty:
            messagebox.showwarning("No data", "No export data available to save.")
            return

        initialdir = ""
        initialfile = ""
        last_file = self.general.get("last_export_csv_filename", "")
        if last_file:
            initialdir = os.path.dirname(last_file)
            initialfile = os.path.basename(last_file)
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            title="Save export as CSV",
            initialdir=initialdir if initialdir else None,
            initialfile=initialfile if initialfile else None,
        )
        if filename:
            try:
                # Only export the selected columns
                self.export_df[selected_cols].to_csv(filename, index=False)
                self.general["last_export_csv_filename"] = filename
                self.save_all_config()
                messagebox.showinfo("Export Successful", f"Exported to {filename}")
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not export CSV:\n{e}")

    def export_excel(self):
        import os
        import sys
        import pandas as pd
        from tkinter import messagebox, filedialog
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        def open_excel_file(filepath):
            try:
                if sys.platform == "win32":
                    os.startfile(filepath)
                elif sys.platform == "darwin":
                    import subprocess
                    subprocess.run(["open", filepath])
                else:
                    import subprocess
                    subprocess.run(["xdg-open", filepath])
            except Exception as e:
                messagebox.showwarning("Open File", f"Could not open file automatically.\n{e}")

        if getattr(self, "export_df", None) is None or self.export_df.empty:
            messagebox.showwarning("No data", "No export data available to save.")
            return

        # Mapping from your DataFrame columns to Excel columns (order matters)
        col_map = {
            'Node Name': 'Node name \nXXXXYYYYZ',
            'filename_dep': 'File name Deployment',
            'filename_rec': 'File name Recovery',
            'Datetime_dep': 'Date - Node Deployment',
            'Datetime_rec': 'Date - Node Recovery'
        }
        export_cols = [c for c in col_map if c in self.export_df.columns]
        if not export_cols:
            messagebox.showwarning("No data", "No matching columns to export.")
            return

        # Prepare DataFrame with mapped columns
        df = self.export_df.loc[:, export_cols].rename(columns=col_map)

        # Ensure date columns are datetime type
        date_cols = [
            'Date - Node Deployment',
            'Date - Node Recovery',
            'Date - Image Node Deployment upload',
            'Date - Image Node Recovery upload'
        ]
        for col in ['Datetime_dep', 'Datetime_rec']:
            if col in self.export_df.columns and col_map[col] in df.columns:
                df[col_map[col]] = pd.to_datetime(df[col_map[col]], errors='coerce').dt.date

        # Two empty columns for upload dates
        df["Date - Image Node Deployment upload"] = ""
        df["Date - Image Node Recovery upload"] = ""

        # "Comment" column is concatenation of DeployedComments and RecoveredComments
        deployed_comment = self.export_df.get("DeployedComments", pd.Series([""]*len(df)))
        recovered_comment = self.export_df.get("RecoveredComments", pd.Series([""]*len(df)))
        df["Comment"] = [
            " | ".join([str(dc) for dc in [d, r] if pd.notnull(dc) and str(dc).strip()])
            for d, r in zip(deployed_comment, recovered_comment)
        ]

        # Set the final column order
        final_cols = [
            'Node name \nXXXXYYYYZ',
            'File name Deployment',
            'File name Recovery',
            'Date - Node Deployment',
            'Date - Node Recovery',
            'Date - Image Node Deployment upload',
            'Date - Image Node Recovery upload',
            'Comment'
        ]
        df = df.loc[:, [col for col in final_cols if col in df.columns]]

        initialdir = ""
        initialfile = ""
        last_file = self.general.get("last_export_excel_filename", "")
        if last_file:
            initialdir = os.path.dirname(last_file)
            initialfile = os.path.basename(last_file)
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            title="Save export as Excel",
            initialdir=initialdir if initialdir else None,
            initialfile=initialfile if initialfile else None,
        )
        if not filename:
            return

        # Export attempt loop (handles file-in-use errors)
        while True:
            try:
                # Export with pandas first
                with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                    df.to_excel(writer, sheet_name="IMAGES", index=False)
                # Now apply formatting
                wb = load_workbook(filename)
                ws = wb["IMAGES"]

                # Set column widths
                widths = [16.14, 55.29, 55.29, 25, 25, 38, 38, 38]
                for i, col in enumerate(ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=ws.max_column)):
                    col_letter = col[0].column_letter
                    width = widths[i] if i < len(widths) else 20  # Default for extra columns
                    ws.column_dimensions[col_letter].width = width

                # Styles
                header_fill = PatternFill("solid", fgColor="DDDDDD")
                cell_font = Font(name="Aptos Narrow", size=11)
                header_font = Font(name="Aptos Narrow", size=11, bold=True)
                center_alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
                left_no_wrap = Alignment(wrap_text=False, vertical="center", horizontal="left")

                # Format header
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment

                # Find indices of columns for further logic
                col_names = [cell.value for cell in ws[1]]
                idx_comment = col_names.index("Comment") if "Comment" in col_names else None

                # Date columns in Excel indices
                date_col_indices = [i for i, name in enumerate(col_names) if name in date_cols]

                # Format all data cells and apply logic (including date formatting)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for colidx, cell in enumerate(row, 0):
                        cell.font = cell_font
                        # "Comment" column: no wrap, left-align; others: wrap, center-align
                        if idx_comment is not None and colidx == idx_comment:
                            cell.alignment = left_no_wrap
                        else:
                            cell.alignment = center_alignment

                        # Date formatting for appropriate columns
                        if colidx in date_col_indices and cell.value:
                            # Try to parse cell value to date if it's a string
                            import datetime
                            from openpyxl.utils.datetime import to_excel

                            try:
                                if isinstance(cell.value, str):
                                    try:
                                        d = pd.to_datetime(cell.value, errors="coerce")
                                        if pd.notnull(d):
                                            cell.value = d.date()
                                    except Exception:
                                        pass
                                if isinstance(cell.value, (datetime.datetime, datetime.date)):
                                    cell.number_format = "yyyy-mm-dd"
                            except Exception:
                                pass

                wb.save(filename)
                self.general["last_export_excel_filename"] = filename
                self.save_all_config()
                messagebox.showinfo("Export Successful", f"Exported to {filename}")
                open_excel_file(filename)
                break
            except PermissionError:
                retry = messagebox.askretrycancel(
                    "File in Use",
                    "Could not export Excel: The file is open in Excel. Please close it and click 'Retry'."
                )
                if not retry:
                    break
            except Exception as e:
                messagebox.showerror("Export Failed", f"Could not export Excel:\n{e}")
                break

    def update_existing_excel_from_ui(self):
        import os
        import sys
        from tkinter import messagebox
        from openpyxl import load_workbook

        # These should be set by your UI
        filename = self.general.get("last_update_excel_file", "")
        col_mapping = self.general.get("last_update_excel_mapping", {})
        if not filename or not os.path.isfile(filename):
            messagebox.showwarning("No file", "Please select a valid Excel file to update.")
            return
        if not col_mapping:
            messagebox.showwarning("No mapping", "No Excel column mapping provided.")
            return
        if getattr(self, "export_df", None) is None or self.export_df.empty:
            messagebox.showwarning("No data", "No export data available to update.")
            return

        self.save_excel_mapping_from_ui()
        col_mapping = self.general.get("last_update_excel_mapping", {})

        try:
            wb = load_workbook(filename)
            ws = wb[wb.sheetnames[0]]
            excel_cols = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            num_data_rows_excel = ws.max_row - 1  # data rows after header
            num_rows_df = len(self.export_df)

            # If more rows in df, add them
            if num_rows_df > num_data_rows_excel:
                rows_to_add = num_rows_df - num_data_rows_excel
                for _ in range(rows_to_add):
                    ws.append([None]*len(excel_cols))  # Add empty row at the bottom

            # Now update all rows
            data_rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(excel_cols)))
            for row_idx in range(num_rows_df):
                for excel_col, df_col in col_mapping.items():
                    if excel_col not in excel_cols or df_col not in self.export_df.columns:
                        continue
                    col_idx = excel_cols.index(excel_col)
                    cell = data_rows[row_idx][col_idx]
                    value = self.export_df.iloc[row_idx][df_col]

                    # Clear cell if value is None, NaN, empty string, or only whitespace
                    if (
                        value is None or
                        (isinstance(value, float) and math.isnan(value)) or
                        (isinstance(value, str) and value.strip() == "")
                    ):
                        cell.value = ""
                    else:
                        cell.value = value

            wb.save(filename)
            messagebox.showinfo("Update successful", f"Excel file updated: {filename}")
            # Open the file as before:
            if sys.platform == "win32":
                os.startfile(filename)
            elif sys.platform == "darwin":
                import subprocess
                subprocess.run(["open", filename])
            else:
                import subprocess
                subprocess.run(["xdg-open", filename])
        except Exception as e:
            messagebox.showerror("Update Failed", f"Could not update Excel file:\n{e}")
        
    def choose_output_dir(self):
        folder = filedialog.askdirectory(title="Select Output Directory")
        if folder:
            self.output_dir_var.set(folder)
            self.config["output_dir"] = folder
            save_config(self.config)

    def load_last_choices(self):
        image_dir = self.general.get("image_dir")
        if image_dir and os.path.isdir(image_dir):
            self.image_dir_var.set(image_dir)
        output_dir = self.general.get("output_dir")
        if output_dir and os.path.isdir(output_dir):
            self.output_dir_var.set(output_dir)
        csv_file = self.general.get("csv_file")
        if csv_file and os.path.isfile(csv_file):
            self.csv_file_var.set(csv_file)

if __name__ == "__main__":
    app = App()
    app.mainloop()
